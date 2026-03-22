"""Excel-Exporter fuer TI-Radar Analyseergebnisse.

Erstellt ein openpyxl-Workbook mit einem Tabellenblatt pro Use-Case.
Header-Zeilen erhalten Formatierung (fett, hellblauer Hintergrund),
Geldwerte werden mit Zahlenformat versehen.

Verwendet dieselben UC-Spaltendefinitionen wie der CSV-Exporter
fuer konsistente Datenstruktur ueber alle Export-Formate.
"""

from __future__ import annotations

import io
from typing import Any

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.exporters.csv_exporter import (
    EXTRA_UC_DEFS,
    UC_COLUMN_DEFS,
    _extract_rows,
)

logger = structlog.get_logger(__name__)


# ---------------------------------------------------------------------------
# Formatierung: Header-Style
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Leicht abgesetztes Blau fuer Zusammenfassungszeilen
SUMMARY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

# Zahlenformat fuer Geldwerte (Euro)
MONEY_FORMAT = '#,##0.00 "EUR"'
PERCENT_FORMAT = "0.00%"
NUMBER_FORMAT = "#,##0"
FLOAT_FORMAT = "0.000000"

# Spalten, die als Geldwerte formatiert werden sollen
MONEY_COLUMNS = {"ec_funding", "funding_amount", "total_funding"}
PERCENT_COLUMNS = {"cagr", "hhi_share", "jaccard_index", "s_curve_r2"}


# ---------------------------------------------------------------------------
# Hauptfunktion
# ---------------------------------------------------------------------------


async def export_excel(data: dict[str, Any], use_cases: list[str]) -> bytes:
    """Erstellt ein Excel-Workbook aus RadarResponse-JSON.

    Erzeugt pro Use-Case ein separates Tabellenblatt mit:
    - Formatierter Header-Zeile (fett, hellblau)
    - Automatisch angepasste Spaltenbreiten
    - Zahlenformatierung fuer Geld- und Prozentwerte
    - Zusammenfassungszeile am Ende (falls sinnvoll)

    Args:
        data: RadarResponse als dict (von Orchestrator oder Cache).
        use_cases: Liste der zu exportierenden UC-Namen.

    Returns:
        Excel-Datei als Bytes (.xlsx-Format).
    """
    wb = Workbook()

    # Standard-Tabellenblatt entfernen (wird durch UC-Blaetter ersetzt)
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    # Uebersichtsblatt erstellen
    _create_overview_sheet(wb, data, use_cases)

    sheets_created = 0

    for uc_name in use_cases:
        panel_data = data.get(uc_name, {})
        if not panel_data:
            logger.debug("uc_panel_leer_excel", uc=uc_name)
            continue

        if uc_name in UC_COLUMN_DEFS:
            display_name, columns, data_key = UC_COLUMN_DEFS[uc_name]
            rows = _extract_rows(panel_data, data_key, columns)
            _create_uc_sheet(wb, display_name, columns, rows)
            sheets_created += 1

        elif uc_name in EXTRA_UC_DEFS:
            display_name, data_key = EXTRA_UC_DEFS[uc_name]
            _create_generic_sheet(wb, display_name, panel_data, data_key)
            sheets_created += 1

        else:
            _create_generic_sheet(wb, uc_name, panel_data, "")
            sheets_created += 1

    # Falls keine Blaetter erstellt wurden, Hinweis-Blatt hinzufuegen
    if sheets_created == 0:
        ws = wb.create_sheet("Keine Daten")
        ws.append(["Keine Daten fuer die angeforderten Use-Cases vorhanden."])

    logger.info("excel_generiert", sheets=sheets_created, use_cases=use_cases)

    # Workbook in Bytes serialisieren
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Uebersichtsblatt
# ---------------------------------------------------------------------------


def _create_overview_sheet(
    wb: Workbook,
    data: dict[str, Any],
    use_cases: list[str],
) -> None:
    """Erstellt ein Uebersichtsblatt mit Metadaten zur Analyse."""
    ws = wb.create_sheet("Uebersicht", 0)

    # Titel
    ws.append(["TI-Radar Export"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    # Metadaten
    technology = data.get("technology", "Unbekannt")
    period = data.get("analysis_period", "")
    total_time = data.get("total_processing_time_ms", 0)
    success_count = data.get("successful_uc_count", 0)
    total_count = data.get("total_uc_count", 12)
    timestamp = data.get("timestamp", "")

    meta_rows = [
        ["Technologie", technology],
        ["Analysezeitraum", period],
        ["Verarbeitungszeit (ms)", total_time],
        ["Erfolgreiche UCs", f"{success_count} / {total_count}"],
        ["Exportierte UCs", ", ".join(use_cases)],
        ["Export-Zeitstempel", timestamp],
    ]

    for row in meta_rows:
        ws.append(row)
        # Label fett
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    # Fehler-Sektion (falls vorhanden)
    uc_errors = data.get("uc_errors", [])
    if uc_errors:
        ws.append([])
        ws.append(["Fehlgeschlagene Use-Cases"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF0000")

        for error in uc_errors:
            if isinstance(error, dict):
                ws.append([
                    error.get("use_case", ""),
                    error.get("error_code", ""),
                    error.get("error_message", ""),
                ])

    # Spaltenbreiten anpassen
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60


# ---------------------------------------------------------------------------
# UC-spezifisches Tabellenblatt
# ---------------------------------------------------------------------------


def _create_uc_sheet(
    wb: Workbook,
    display_name: str,
    columns: list[str],
    rows: list[list[Any]],
) -> None:
    """Erstellt ein Tabellenblatt fuer einen spezifischen Use-Case.

    Features:
    - Formatierte Header-Zeile (fett, blauer Hintergrund, weisse Schrift)
    - Spaltenbreiten automatisch an Inhalt angepasst
    - Zahlenformate fuer Geld- und Prozentwerte
    - Autofilter auf der Header-Zeile
    """
    # Blattnamen auf max. 31 Zeichen kuerzen (Excel-Limit)
    sheet_name = display_name[:31]
    ws = wb.create_sheet(sheet_name)

    # Header-Zeile schreiben und formatieren
    ws.append(columns)
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    # Datenzeilen schreiben
    for row in rows:
        formatted_row = []
        for col_idx, val in enumerate(row):
            if val is None or val == "":
                formatted_row.append("")
            elif isinstance(val, (int, float)):
                formatted_row.append(val)
            else:
                formatted_row.append(str(val))
        ws.append(formatted_row)

    # Zahlenformate anwenden
    _apply_number_formats(ws, columns)

    # Spaltenbreiten automatisch anpassen
    _auto_column_widths(ws, columns)

    # Autofilter aktivieren
    if rows:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    logger.debug(
        "excel_blatt_erstellt",
        sheet=sheet_name,
        rows=len(rows),
        columns=len(columns),
    )


# ---------------------------------------------------------------------------
# Generisches Tabellenblatt
# ---------------------------------------------------------------------------


def _create_generic_sheet(
    wb: Workbook,
    display_name: str,
    panel_data: dict[str, Any],
    data_key: str,
) -> None:
    """Erstellt ein generisches Tabellenblatt fuer UCs ohne spezifische Definition.

    Versucht die Daten als Tabelle darzustellen. Fallback auf Key-Value-Format.
    """
    sheet_name = display_name[:31]
    ws = wb.create_sheet(sheet_name)

    # Daten unter dem angegebenen Schluessel suchen
    items = panel_data.get(data_key, []) if data_key else []
    if not items and isinstance(panel_data.get("data"), dict):
        items = panel_data["data"].get(data_key, []) if data_key else []

    if isinstance(items, list) and items and isinstance(items[0], dict):
        # Tabellenformat: Spaltenkoepfe aus erstem Eintrag
        columns = list(items[0].keys())

        # Header schreiben
        ws.append(columns)
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT

        # Datenzeilen
        for item in items:
            row = []
            for col in columns:
                val = item.get(col, "")
                if isinstance(val, (dict, list)):
                    row.append(str(val)[:200])
                elif val is None:
                    row.append("")
                else:
                    row.append(val)
            ws.append(row)

        _auto_column_widths(ws, columns)

        # Autofilter
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(items) + 1}"

    else:
        # Key-Value-Format
        columns = ["Schluessel", "Wert"]
        ws.append(columns)
        for col_idx in range(1, 3):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT

        for key, value in panel_data.items():
            if key in ("metadata", "data", "result"):
                continue
            if isinstance(value, (dict, list)):
                ws.append([str(key), str(value)[:500]])
            else:
                ws.append([str(key), str(value) if value is not None else ""])

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60


# ---------------------------------------------------------------------------
# Formatierungs-Hilfsfunktionen
# ---------------------------------------------------------------------------


def _apply_number_formats(ws: Any, columns: list[str]) -> None:
    """Wendet Zahlenformate auf Spalten an, basierend auf dem Spaltennamen.

    Geldwerte (ec_funding etc.) erhalten EUR-Format,
    Prozentwerte (cagr, hhi_share) erhalten Prozent-Format.
    """
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)

        if col_name in MONEY_COLUMNS:
            fmt = MONEY_FORMAT
        elif col_name in PERCENT_COLUMNS:
            fmt = FLOAT_FORMAT  # Werte kommen als Dezimalzahlen, nicht 0-1
        elif col_name in ("patent_count", "project_count", "citations",
                          "co_occurrence_count", "persistence_years"):
            fmt = NUMBER_FORMAT
        else:
            continue

        # Format auf alle Datenzeilen anwenden (ab Zeile 2)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and cell.value != "":
                cell.number_format = fmt


def _auto_column_widths(ws: Any, columns: list[str]) -> None:
    """Passt Spaltenbreiten automatisch an den Inhalt an.

    Beruecksichtigt sowohl Header-Laenge als auch den laengsten Wert
    in den ersten 100 Zeilen (Performance-Limit bei grossen Datensaetzen).
    """
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)

        # Mindestbreite: Header-Laenge + Padding
        max_width = len(col_name) + 4

        # Maximale Breite aus den ersten 100 Datenzeilen
        max_rows_to_check = min(ws.max_row, 101)
        for row_idx in range(2, max_rows_to_check + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell_len = len(str(cell.value))
                max_width = max(max_width, min(cell_len + 2, 50))

        ws.column_dimensions[col_letter].width = max_width
